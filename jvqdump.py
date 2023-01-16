#!/usr/bin/env python3

"""Dump data in an iNES ROM file.

Usage:
  poetry install
  poetry run ./jvqdump.py input_ines_file_path output_excel_file_path
"""

import argparse
import binascii
import collections
import dataclasses
import enum
import itertools
import operator
import struct
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

import openpyxl
import openpyxl.styles

_ENEMY_GROUP_PATTERN_LIST_ID_COUNT = 155
_MAP_ID_COUNT = 350
_MAX_ACTIONS_PER_ENEMY = 8
_ENEMY_ID_COUNT = 173


def _read_prg_rom(input_ines_file_path: str) -> bytes:
    ines_header_byte_size = 16
    with open(input_ines_file_path, "rb") as f:
        header_bytes = f.read(ines_header_byte_size)
        if len(header_bytes) != ines_header_byte_size:
            raise ValueError("Invalid iNES header")
        if header_bytes[0:4] != b"NES\x1a":
            raise ValueError("Invalid signature")
        prg_rom_byte_size = header_bytes[4] * 0x4000
        prg_rom_bytes = f.read(prg_rom_byte_size)
        if len(prg_rom_bytes) != prg_rom_byte_size:
            raise ValueError("Insufficient PRG ROM")
        return prg_rom_bytes


def _decode_string(str_bytes: bytes) -> str:
    byte_to_str = {
        # 0x00: ??
        0x01: "あ",
        0x02: "い",
        0x03: "う",
        0x04: "え",
        0x05: "お",
        0x06: "か",
        0x07: "き",
        0x08: "く",
        0x09: "け",
        0x0A: "こ",
        0x0B: "さ",
        0x0C: "し",
        0x0D: "す",
        0x0E: "せ",
        0x0F: "そ",
        # 0x10: 半濁点
        0x11: "た",
        0x12: "ち",
        0x13: "つ",
        0x14: "て",
        0x15: "と",
        0x16: "な",
        0x17: "に",
        0x18: "ぬ",
        0x19: "ね",
        0x1A: "の",
        0x1B: "は",
        0x1C: "ひ",
        0x1D: "ふ",
        0x1E: "へ",
        0x1F: "ほ",
        0x20: "両",
        0x21: "ま",
        0x22: "み",
        0x23: "む",
        0x24: "め",
        0x25: "も",
        0x26: "や",
        0x27: "ゆ",
        0x28: "よ",
        0x29: "ら",
        0x2A: "り",
        0x2B: "る",
        0x2C: "れ",
        0x2D: "ろ",
        0x2E: "わ",
        0x2F: "を",
        0x30: "０",
        0x31: "１",
        0x32: "２",
        0x33: "３",
        0x34: "４",
        0x35: "５",
        0x36: "６",
        0x37: "７",
        0x38: "８",
        0x39: "９",
        0x3A: "！",
        0x3B: "？",
        0x3C: "「",
        0x3D: "／",
        0x3E: "・",
        0x3F: "ん",
        0x40: "■",
        0x41: "ア",
        0x42: "イ",
        0x43: "ウ",
        0x44: "エ",
        0x45: "オ",
        0x46: "カ",
        0x47: "キ",
        0x48: "ク",
        0x49: "ケ",
        0x4A: "コ",
        0x4B: "サ",
        0x4C: "シ",
        0x4D: "ス",
        0x4E: "セ",
        0x4F: "ソ",
        # 0x50: ??
        0x51: "タ",
        0x52: "チ",
        0x53: "ツ",
        0x54: "テ",
        0x55: "ト",
        0x56: "ナ",
        0x57: "ニ",
        0x58: "ヌ",
        0x59: "ネ",
        0x5A: "ノ",
        0x5B: "ハ",
        0x5C: "ヒ",
        0x5D: "フ",
        0x5E: "ー",
        0x5F: "ホ",
        # 0x60: ??
        0x61: "マ",
        0x62: "ミ",
        0x63: "ム",
        0x64: "メ",
        0x65: "モ",
        0x66: "ヤ",
        0x67: "ユ",
        0x68: "ヨ",
        0x69: "ラ",
        0x6A: "◆",
        0x6B: "ル",
        0x6C: "レ",
        0x6D: "ロ",
        0x6E: "ワ",
        0x6F: "超",
        # 0x70: Frame
        # 0x71: Frame
        # 0x72: Frame
        # 0x73: Frame
        # 0x74: Frame
        # 0x75: Frame
        # 0x76: Frame
        # 0x77: Frame
        0x78: "ッ",
        0x79: "ャ",
        0x7A: "ュ",
        0x7B: "ョ",
        0x7C: "ァ",
        0x7D: "っ",
        0x7E: "ゃ",
        0x7F: "ン",
        0x80: "ゅ",
        0x81: "ょ",
        # 0x82: ??
        0x83: "ェ",
        # 0x84: ??
        # 0x85: ??
        0x86: "が",
        0x87: "ぎ",
        0x88: "ぐ",
        0x89: "げ",
        0x8A: "ご",
        0x8B: "ざ",
        0x8C: "じ",
        0x8D: "ず",
        0x8E: "ぜ",
        0x8F: "ぞ",
        # 0x90: Invalid
        0x91: "だ",
        0x92: "ぢ",
        0x93: "づ",
        0x94: "で",
        0x95: "ど",
        # 0x96: Invalid
        # 0x97: Invalid
        # 0x98: Invalid
        # 0x99: Invalid
        # 0x9A: Invalid
        0x9B: "ば",
        0x9C: "び",
        0x9D: "ぶ",
        0x9E: "べ",
        0x9F: "ぼ",
        # 0xA0: Invalid
        0xA1: "ぱ",
        0xA2: "ぴ",
        0xA3: "ぶ",
        0xA4: "ぺ",
        0xA5: "ぽ",
        0xA6: "パ",
        0xA7: "ピ",
        0xA8: "プ",
        # 0xA9: Invalid
        0xAA: "ポ",
        # 0xAB: Invalid
        # 0xAC: Invalid
        # 0xAD: Invalid
        # 0xAE: Invalid
        # 0xAF: Invalid
        # 0xB0: Invalid
        # 0xB1: Invalid
        # 0xB2: Invalid
        # 0xB3: Invalid
        # 0xB4: Invalid
        # 0xB5: Invalid
        # 0xB6: Invalid
        # 0xB7: Invalid
        # 0xB8: Invalid
        # 0xB9: Invalid
        # 0xBA: Invalid
        # 0xBB: Invalid
        # 0xBC: Invalid
        # 0xBD: Invalid
        # 0xBE: Invalid
        # 0xBF: Invalid
        # 0xC0: Invalid
        # 0xC1: Invalid
        # 0xC2: Invalid
        # 0xC3: Invalid
        # 0xC4: Invalid
        # 0xC5: Invalid
        0xC6: "ガ",
        0xC7: "ギ",
        0xC8: "グ",
        0xC9: "ゲ",
        0xCA: "ゴ",
        0xCB: "ザ",
        0xCC: "ジ",
        0xCD: "ズ",
        0xCE: "ゼ",
        0xCF: "ゾ",
        # 0xD0: Invalid
        0xD1: "ダ",
        0xD2: "ヂ",
        0xD3: "ヅ",
        0xD4: "デ",
        0xD5: "ド",
        # 0xD6: Invalid
        # 0xD7: Invalid
        # 0xD8: Invalid
        # 0xD9: Invalid
        # 0xDA: Invalid
        0xDB: "バ",
        0xDC: "ビ",
        0xDD: "ブ",
        # 0xDE: Invalid
        0xDF: "ボ",
        # 0xE0: Invalid
        # 0xE1: Invalid
        # 0xE2: Invalid
        # 0xE3: Invalid
        # 0xE4: Invalid
        # 0xE5: Invalid
        # 0xE6: Invalid
        # 0xE7: Invalid
        # 0xE8: Invalid
        # 0xE9: Invalid
        # 0xEA: Invalid
        # 0xEB: Invalid
        # 0xEC: Invalid
        # 0xED: Invalid
        # 0xEE: Invalid
        # 0xEF: Invalid
        # 0xF0: Invalid
        # 0xF1: Invalid
        # 0xF2: Invalid
        # 0xF3: Invalid
        # 0xF4: Invalid
        # 0xF5: Invalid
        # 0xF6: Invalid
        # 0xF7: Invalid
        # 0xF8: Invalid
        # 0xF9: Invalid
        # 0xFA: Invalid
        # 0xFB: Invalid
        # 0xFC: Invalid
        # 0xFD: Invalid
        # 0xFE: Invalid
        0xFF: "　",  # Space
    }
    return "".join(byte_to_str[s] for s in str_bytes)


@enum.unique
class _PlayerCharacterType(enum.Enum):
    JUVEI = "じゅうべえ"
    RYUHIME = "りゅうひめ"
    WOLF = "ウルフ・シロ"
    IWAN = "イワン・ガンちゃん"
    ONITAN = "オニタン"
    HINOTORI = "ひのとり"
    SARUBOSS = "サルボス"
    PENTA = "ペンタ"
    LUCKY = "ラッキー"


@dataclasses.dataclass(frozen=True)
class _PlayerCharacterLevel:
    level: int
    hp: int  # 命
    cp: int  # 超力
    attack: int  # 攻撃
    defense: int  # 守備
    speed: int  # スピード
    ten: int  # 天の守り
    shin: int  # 芯の強さ
    atama: int  # 頭の良さ
    experience_required_from_previous_level: int
    accumulated_experience_required: int


@dataclasses.dataclass(frozen=True)
class _PlayerCharacter:
    player_character_type: _PlayerCharacterType
    max_level: int
    levels: Sequence[_PlayerCharacterLevel]


def _get_player_character_max_level(prg_rom_bytes: bytes, player_character_type: _PlayerCharacterType) -> int:
    prg_rom_address = {
        _PlayerCharacterType.JUVEI: 0x03C097,
        _PlayerCharacterType.WOLF: 0x03C09C,
        _PlayerCharacterType.RYUHIME: 0x03C098,
        _PlayerCharacterType.IWAN: 0x03C09A,
        _PlayerCharacterType.ONITAN: 0x03C09B,
        _PlayerCharacterType.HINOTORI: 0x03C09D,
        _PlayerCharacterType.SARUBOSS: 0x03C09E,
        _PlayerCharacterType.PENTA: 0x03C09F,
        _PlayerCharacterType.LUCKY: 0x03C0A0,
    }[player_character_type]
    return prg_rom_bytes[prg_rom_address]


def _get_player_character(prg_rom_bytes: bytes, player_character_type: _PlayerCharacterType) -> _PlayerCharacter:
    status_start_prg_rom_address = {
        _PlayerCharacterType.JUVEI: 0x027713,
        _PlayerCharacterType.WOLF: 0x0278D5,
        _PlayerCharacterType.RYUHIME: 0x027A97,
        _PlayerCharacterType.IWAN: 0x027C59,
        _PlayerCharacterType.ONITAN: 0x027DF0,
        _PlayerCharacterType.HINOTORI: 0x027E4A,
        _PlayerCharacterType.SARUBOSS: 0x027EA4,
        _PlayerCharacterType.PENTA: 0x027EEC,
        _PlayerCharacterType.LUCKY: 0x027F46,
    }[player_character_type]
    experience_start_prg_rom_address = {
        _PlayerCharacterType.JUVEI: 0x027519,
        _PlayerCharacterType.WOLF: 0x02757D,
        _PlayerCharacterType.RYUHIME: 0x0275E1,
        _PlayerCharacterType.IWAN: 0x027645,
        _PlayerCharacterType.ONITAN: 0x002769F,
        _PlayerCharacterType.HINOTORI: 0x0276B3,
        _PlayerCharacterType.SARUBOSS: 0x0276C7,
        _PlayerCharacterType.PENTA: 0x0276D7,
        _PlayerCharacterType.LUCKY: 0x0276EB,
    }[player_character_type]
    max_level = _get_player_character_max_level(prg_rom_bytes, player_character_type)
    experience_unit_byte_size = 2
    status_unit_byte_size = 9
    accumulated_experience_required = 0
    player_character_levels = []
    for level in range(1, max_level + 1):
        status_offset = status_start_prg_rom_address + (level - 1) * status_unit_byte_size
        (hp, cp, attack, defense, speed, ten, shin, atama) = struct.unpack_from("<HBBBBBBB", prg_rom_bytes, status_offset)
        experience_offset = experience_start_prg_rom_address + (level - 1) * experience_unit_byte_size
        (experience_required,) = struct.unpack_from("<H", prg_rom_bytes, experience_offset)
        accumulated_experience_required += experience_required
        player_character_level = _PlayerCharacterLevel(
            level,
            hp,
            cp,
            attack,
            defense,
            speed,
            ten,
            shin,
            atama,
            experience_required,
            accumulated_experience_required,
        )
        player_character_levels.append(player_character_level)
    return _PlayerCharacter(player_character_type, max_level, tuple(player_character_levels))


@dataclasses.dataclass(frozen=True)
class _EnemyGroupPatternList:
    enemy_group_pattern_ids: Sequence[int]


def _get_enemy_group_pattern_list(prg_rom_bytes: bytes, enemy_group_pattern_list_id: int) -> _EnemyGroupPatternList:
    prg_rom_address = 0x00F704
    list_id = 0
    while list_id < enemy_group_pattern_list_id:
        if prg_rom_bytes[prg_rom_address] == 0xFF:
            list_id += 1
        prg_rom_address += 1
    enemy_group_pattern_ids = []
    while prg_rom_bytes[prg_rom_address] != 0xFF:
        enemy_group_pattern_ids.append(prg_rom_bytes[prg_rom_address])
        prg_rom_address += 1
    return _EnemyGroupPatternList(enemy_group_pattern_ids=tuple(enemy_group_pattern_ids))


@dataclasses.dataclass(frozen=True)
class _EnemyGroupPattern:
    enemy_group_size: int
    enemy_group_0_size: int
    enemy_group_0_enemy_id: Optional[int]
    enemy_group_1_size: int
    enemy_group_1_enemy_id: Optional[int]
    enemy_group_2_size: int
    enemy_group_2_enemy_id: Optional[int]


def _get_enemy_group_pattern(prg_rom_bytes: bytes, enemy_group_pattern_id: int) -> _EnemyGroupPattern:
    assert 0 < enemy_group_pattern_id
    prg_rom_address = 0x00FA48
    pattern_id = 0
    while pattern_id < enemy_group_pattern_id:
        encoded_enemy_group_size = prg_rom_bytes[prg_rom_address]
        prg_rom_address += 1
        enemy_group_0_size = (encoded_enemy_group_size & 0xE0) >> 5
        if enemy_group_0_size != 0:
            prg_rom_address += 1
        enemy_group_1_size = (encoded_enemy_group_size & 0x1C) >> 2
        if enemy_group_1_size != 0:
            prg_rom_address += 1
        enemy_group_2_size = encoded_enemy_group_size & 0x03
        if enemy_group_2_size != 0:
            prg_rom_address += 1
        pattern_id += 1
    encoded_enemy_group_size = prg_rom_bytes[prg_rom_address]
    enemy_group_size = 0
    prg_rom_address += 1
    enemy_group_0_size = (encoded_enemy_group_size & 0xE0) >> 5
    enemy_group_0_enemy_id = None
    if enemy_group_0_size != 0:
        enemy_group_0_enemy_id = prg_rom_bytes[prg_rom_address]
        enemy_group_size += 1
        prg_rom_address += 1
    enemy_group_1_size = (encoded_enemy_group_size & 0x1C) >> 2
    enemy_group_1_enemy_id = None
    if enemy_group_1_size != 0:
        enemy_group_1_enemy_id = prg_rom_bytes[prg_rom_address]
        enemy_group_size += 1
        prg_rom_address += 1
    enemy_group_2_size = encoded_enemy_group_size & 0x03
    enemy_group_2_enemy_id = None
    if enemy_group_2_size != 0:
        enemy_group_2_enemy_id = prg_rom_bytes[prg_rom_address]
        enemy_group_size += 1
        prg_rom_address += 1
    return _EnemyGroupPattern(
        enemy_group_size=enemy_group_size,
        enemy_group_0_size=enemy_group_0_size,
        enemy_group_0_enemy_id=enemy_group_0_enemy_id,
        enemy_group_1_size=enemy_group_1_size,
        enemy_group_1_enemy_id=enemy_group_1_enemy_id,
        enemy_group_2_size=enemy_group_2_size,
        enemy_group_2_enemy_id=enemy_group_2_enemy_id,
    )


@dataclasses.dataclass(frozen=True)
class _EnemyActionId:
    action_ids: Sequence[int]


def _get_enemy_action_id(prg_rom_bytes: bytes, raw_action_id: int) -> _EnemyActionId:
    assert 1 <= raw_action_id
    start_prg_rom_address = 0x00FF93 + (raw_action_id - 1)
    action_id = prg_rom_bytes[start_prg_rom_address]
    if action_id < 0x80:
        return _EnemyActionId(action_ids=tuple([action_id]))
    action_id = (action_id << 1) & 0xFF
    return _EnemyActionId(
        action_ids=tuple(
            [
                prg_rom_bytes[0x00FFEE + action_id],
                prg_rom_bytes[0x00FFEE + action_id + 1],
            ]
        )
    )


@dataclasses.dataclass(frozen=True)
class _EnemyActionPattern:
    raw_action_id_0: int
    raw_action_id_1: int
    raw_action_id_2: int
    raw_action_id_3: int
    action_id_0: _EnemyActionId
    action_id_1: _EnemyActionId
    action_id_2: _EnemyActionId
    action_id_3: _EnemyActionId
    action_threshold_0: int
    action_threshold_1: int
    action_threshold_2: int
    action_threshold_3: int


def _get_enemy_action_pattern(prg_rom_bytes: bytes, enemy_action_pattern_id: int) -> _EnemyActionPattern:
    assert enemy_action_pattern_id <= 0x3F
    enemy_action_pattern_unit_byte_size = 8
    start_prg_rom_address = 0x00F442 + enemy_action_pattern_id * enemy_action_pattern_unit_byte_size
    enemy_action_pattern = _EnemyActionPattern(
        raw_action_id_0=prg_rom_bytes[start_prg_rom_address],
        raw_action_id_1=prg_rom_bytes[start_prg_rom_address + 1],
        raw_action_id_2=prg_rom_bytes[start_prg_rom_address + 2],
        raw_action_id_3=prg_rom_bytes[start_prg_rom_address + 3],
        action_id_0=_get_enemy_action_id(prg_rom_bytes, prg_rom_bytes[start_prg_rom_address]),
        action_id_1=_get_enemy_action_id(prg_rom_bytes, prg_rom_bytes[start_prg_rom_address + 1]),
        action_id_2=_get_enemy_action_id(prg_rom_bytes, prg_rom_bytes[start_prg_rom_address + 2]),
        action_id_3=_get_enemy_action_id(prg_rom_bytes, prg_rom_bytes[start_prg_rom_address + 3]),
        action_threshold_0=prg_rom_bytes[start_prg_rom_address + 4],
        action_threshold_1=prg_rom_bytes[start_prg_rom_address + 5],
        action_threshold_2=prg_rom_bytes[start_prg_rom_address + 6],
        action_threshold_3=prg_rom_bytes[start_prg_rom_address + 7],
    )
    # Validate that the sum of the thresholds should be around 0xFF.
    assert 0xFA <= (enemy_action_pattern.action_threshold_0 + enemy_action_pattern.action_threshold_1 + enemy_action_pattern.action_threshold_2 + enemy_action_pattern.action_threshold_3) <= 0x100
    return enemy_action_pattern


def _aggregate_enemy_action_pattern(enemy_action_pattern: _EnemyActionPattern) -> Dict[int, int]:
    action_threshold_0 = enemy_action_pattern.action_threshold_0
    action_threshold_1 = enemy_action_pattern.action_threshold_1
    action_threshold_2 = enemy_action_pattern.action_threshold_2
    action_threshold_3 = 0x100 - (action_threshold_0 + action_threshold_1 + action_threshold_2)
    assert 0x00 <= action_threshold_3
    threshold_by_action_id: Dict[int, int] = collections.defaultdict(int)
    for action_id in enemy_action_pattern.action_id_0.action_ids:
        threshold_by_action_id[action_id] += action_threshold_0 * 2 // len(enemy_action_pattern.action_id_0.action_ids)
    for action_id in enemy_action_pattern.action_id_1.action_ids:
        threshold_by_action_id[action_id] += action_threshold_1 * 2 // len(enemy_action_pattern.action_id_1.action_ids)
    for action_id in enemy_action_pattern.action_id_2.action_ids:
        threshold_by_action_id[action_id] += action_threshold_2 * 2 // len(enemy_action_pattern.action_id_2.action_ids)
    for action_id in enemy_action_pattern.action_id_3.action_ids:
        threshold_by_action_id[action_id] += action_threshold_3 * 2 // len(enemy_action_pattern.action_id_3.action_ids)
    assert sum(threshold_by_action_id.values()) == 0x100 * 2, sum(threshold_by_action_id.values())
    return threshold_by_action_id


@dataclasses.dataclass(frozen=True)
class _ActionName:
    action_name_bytes: bytes
    action_name: str


def _get_action_name(prg_rom_bytes: bytes, action_id: int) -> _ActionName:
    if action_id == 0xEE:
        return _ActionName(
            action_name_bytes=b"",
            action_name="こうげき　　",
        )
    action_name_unit_byte_size = 8
    start_prg_rom_address = 0x02634D + action_id * action_name_unit_byte_size
    action_name_bytes = prg_rom_bytes[start_prg_rom_address : start_prg_rom_address + action_name_unit_byte_size - 1]
    return _ActionName(
        action_name_bytes=action_name_bytes,
        action_name=_decode_string(action_name_bytes),
    )


@dataclasses.dataclass(frozen=True)
class _ItemDropPattern:
    drop_item_id_0: int
    drop_item_id_1: int
    item_drop_threshold: int


def _get_item_drop_pattern(prg_rom_bytes: bytes, item_drop_pattern_id: int) -> _ItemDropPattern:
    assert item_drop_pattern_id <= 0x3F
    item_drop_pattern_unit_byte_size = 3
    start_prg_rom_address = 0xF643 + item_drop_pattern_id * item_drop_pattern_unit_byte_size
    return _ItemDropPattern(
        drop_item_id_0=prg_rom_bytes[start_prg_rom_address],
        drop_item_id_1=prg_rom_bytes[start_prg_rom_address + 1],
        item_drop_threshold=prg_rom_bytes[start_prg_rom_address + 2],
    )


@dataclasses.dataclass(frozen=True)
class _ItemName:
    item_name_bytes: bytes
    item_name: str


def _get_item_name(prg_rom_bytes: bytes, item_id: int) -> _ItemName:
    assert 0x00 < item_id
    item_name_unit_byte_size = 8
    start_prg_rom_address = 0x026685 + item_id * item_name_unit_byte_size
    item_name_bytes = prg_rom_bytes[start_prg_rom_address : start_prg_rom_address + item_name_unit_byte_size - 1]
    return _ItemName(
        item_name_bytes=item_name_bytes,
        item_name=_decode_string(item_name_bytes),
    )


@dataclasses.dataclass(frozen=True)
class _EnemyName:
    enemy_name_bytes: bytes
    enemy_name: str


def _get_enemy_name(prg_rom_bytes: bytes, enemy_id: int) -> _EnemyName:
    assert 0x00 < enemy_id
    enemy_name_unit_byte_size = 8
    start_prg_rom_address = 0x25AD5 + enemy_id * enemy_name_unit_byte_size
    enemy_name_bytes = prg_rom_bytes[start_prg_rom_address : start_prg_rom_address + enemy_name_unit_byte_size - 1]
    return _EnemyName(
        enemy_name_bytes=enemy_name_bytes,
        enemy_name=_decode_string(enemy_name_bytes),
    )


@dataclasses.dataclass(frozen=True)
class _Enemy:
    hp: int
    min_hp: int
    max_hp: int
    cp: int
    attack: int
    defense: int
    speed: int
    money: int
    experience: int
    escapable: bool
    attack_twice: bool
    hittability: int
    mihagito_endurance: bool
    kurusu_endurance: bool
    beto_endurance: bool
    choriki_endurance_1: int
    choriki_endurance_2: int
    choriki_endurance_3_4: int
    choriki_endurance_5: int
    choriki_endurance_6: int
    mahuuji_endurance: int
    mahuuji_effectiveness: int
    lullaby_endurance: int
    lullaby_effectiveness: int
    parapa_endurance: int
    parapa_effectiveness: int
    action_pattern_id: int
    item_drop_pattern_id: int


def _decode_enemy_value(enemy_data: Dict[int, int], y: int) -> int:
    assert 0 <= y
    value = ((enemy_data[0x6000] >> y) * 0x100 + enemy_data[0x6001 + y]) & 0x1FF
    bit_count = 0
    for i in range(5):
        ends_with_one = value & 0x01 != 0
        if ends_with_one:
            bit_count += 1
        value = (value >> 1) & 0xFF
        if not ends_with_one:
            break
    if bit_count == 0:
        return value
    elif bit_count <= 3:
        return (value + 1) * (10**bit_count)
    elif bit_count == 4:
        v = (value + 1) * 1000
        if (v & 0x6000) != 0:
            return 0xFFFF
        else:
            # エイりアンドー
            return 0x9C50
    raise ValueError("Invalid enemy_data.")


def _calculate_enemy_hp_range(hp: int) -> Tuple[int, int]:
    diff = min(hp // 8, 0xFF)
    return (hp - diff, hp + diff)


def _get_enemy(prg_rom_bytes: bytes, enemy_id: int) -> _Enemy:
    assert 0 < enemy_id
    enemy_status_unit_byte_size = 0x14
    enemy_status_start_prg_rom_address = 0x00E1C2 + (enemy_id - 1) * enemy_status_unit_byte_size
    enemy_data: Dict[int, int] = {}
    for i in range(enemy_status_unit_byte_size):
        enemy_data[0x6000 + i] = prg_rom_bytes[enemy_status_start_prg_rom_address + i]
    experience = _decode_enemy_value(enemy_data, 0)
    hp = _decode_enemy_value(enemy_data, 1)
    (min_hp, max_hp) = _calculate_enemy_hp_range(hp)
    attack = _decode_enemy_value(enemy_data, 2)
    defense = _decode_enemy_value(enemy_data, 3)
    cp = _decode_enemy_value(enemy_data, 4)
    speed = _decode_enemy_value(enemy_data, 5)
    money = _decode_enemy_value(enemy_data, 6)
    escapable = (enemy_data[0x6008] & 0x04) == 0
    attack_twice = (enemy_data[0x6008] & 0x30) in (0x10, 0x20)
    hittability = (enemy_data[0x6008] & 0xC0) >> 6
    mihagito_endurance = (enemy_data[0x6008] & 0x08) != 0
    kurusu_endurance = not escapable
    beto_endurance = (enemy_data[0x6008] & 0x02) != 0
    # #$C0:属性2(火炎系)
    choriki_endurance_2 = (enemy_data[0x6009] & 0xC0) >> 6
    # #$C0:属性5(電撃系), #$30:属性3-4(水撃系・氷結系), #$0C:属性1(地震系), #$03:属性6(爆発系)
    choriki_endurance_5 = (enemy_data[0x600A] & 0xC0) >> 6
    choriki_endurance_3_4 = (enemy_data[0x600A] & 0x30) >> 4
    choriki_endurance_1 = (enemy_data[0x600A] & 0x0C) >> 2
    choriki_endurance_6 = enemy_data[0x600A] & 0x03
    mahuuji_endurance = (enemy_data[0x600B] & 0xC0) >> 6
    mahuuji_effectiveness = (enemy_data[0x600B] & 0x30) >> 4
    lullaby_endurance = (enemy_data[0x600B] & 0x0C) >> 2
    lullaby_effectiveness = enemy_data[0x600B] & 0x03
    parapa_endurance = (enemy_data[0x600C] & 0x0C) >> 2
    parapa_effectiveness = enemy_data[0x600C] & 0x03
    action_pattern_id = enemy_data[0x6009] & 0x3F
    item_drop_pattern_id = enemy_data[0x600D] & 0x3F
    return _Enemy(
        hp=hp,
        min_hp=min_hp,
        max_hp=max_hp,
        cp=cp,
        attack=attack,
        defense=defense,
        speed=speed,
        experience=experience,
        money=money,
        escapable=escapable,
        attack_twice=attack_twice,
        hittability=hittability,
        mihagito_endurance=mihagito_endurance,
        kurusu_endurance=kurusu_endurance,
        beto_endurance=beto_endurance,
        choriki_endurance_1=choriki_endurance_1,
        choriki_endurance_2=choriki_endurance_2,
        choriki_endurance_3_4=choriki_endurance_3_4,
        choriki_endurance_5=choriki_endurance_5,
        choriki_endurance_6=choriki_endurance_6,
        mahuuji_endurance=mahuuji_endurance,
        mahuuji_effectiveness=mahuuji_effectiveness,
        lullaby_endurance=lullaby_endurance,
        lullaby_effectiveness=lullaby_effectiveness,
        parapa_endurance=parapa_endurance,
        parapa_effectiveness=parapa_effectiveness,
        action_pattern_id=action_pattern_id,
        item_drop_pattern_id=item_drop_pattern_id,
    )


@dataclasses.dataclass(frozen=True)
class _Map:
    enemy_group_pattern_list_id: int
    encounter_threshold_id: int


def _get_map(prg_rom_bytes: bytes, map_id: int) -> _Map:
    map_unit_byte_size = 12
    start_prg_rom_address = 0x8004 + map_id * map_unit_byte_size
    v008F = prg_rom_bytes[start_prg_rom_address + 1]
    if v008F == 0x01:
        enemy_group_pattern_list_id = 0x78
    elif v008F == 0x02:
        enemy_group_pattern_list_id = 0x00
    else:
        v0091 = prg_rom_bytes[start_prg_rom_address + 3]
        enemy_group_pattern_list_id = v0091
    if enemy_group_pattern_list_id >= 0x9B:
        enemy_group_pattern_list_id = 0x01
    v0098 = prg_rom_bytes[start_prg_rom_address + 10]
    encounter_threshold_id = v0098 >> 5
    return _Map(enemy_group_pattern_list_id=enemy_group_pattern_list_id, encounter_threshold_id=encounter_threshold_id)


_LEFT_ALIGNMENT = openpyxl.styles.Alignment(horizontal="left", vertical="top")
_RIGHT_ALIGNMENT = openpyxl.styles.Alignment(horizontal="right", vertical="top")


@dataclasses.dataclass(frozen=True)
class _ColumnStyle:
    caption: str
    cell_alignment: openpyxl.styles.Alignment = _LEFT_ALIGNMENT


def _fill_worksheet_header_row(worksheet: openpyxl.worksheet.worksheet.Worksheet, row_index: int, header_row: List[_ColumnStyle]) -> None:
    for column_index, header in enumerate(header_row, 1):
        cell = worksheet.cell(column=column_index, row=row_index)
        cell.value = header.caption


def _fill_worksheet_row(worksheet: openpyxl.worksheet.worksheet.Worksheet, row_index: int, header_row: List[_ColumnStyle], row: List[Any]) -> None:
    for column_index, (header_column, column) in enumerate(zip(header_row, row), 1):
        cell = worksheet.cell(column=column_index, row=row_index)
        cell.value = column
        cell.alignment = header_column.cell_alignment


def _fill_player_character(prg_rom_bytes: bytes, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    row_index = 1
    header_row = [
        _ColumnStyle(caption="キャラクター名"),
        _ColumnStyle(caption="最大レベル", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="レベル", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="最大命", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="最大超力", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="攻撃", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="守備", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="スピード", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="天の守り", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="芯の強さ", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="頭の良さ", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="前レベルからの必要経験値", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="累積必要経験値", cell_alignment=_RIGHT_ALIGNMENT),
    ]
    _fill_worksheet_header_row(worksheet, row_index, header_row)
    worksheet.freeze_panes = "B2"
    for player_character_type in _PlayerCharacterType:
        player_character = _get_player_character(prg_rom_bytes, player_character_type)
        for level in player_character.levels:
            row = [
                player_character.player_character_type.value,
                player_character.max_level,
                level.level,
                level.hp,
                level.cp,
                level.attack,
                level.defense,
                level.speed,
                level.ten,
                level.shin,
                level.atama,
                level.experience_required_from_previous_level,
                level.accumulated_experience_required,
            ]
            row_index += 1
            _fill_worksheet_row(worksheet, row_index, header_row, row)


def _get_chapter_name_of_enemy(enemy_id: int) -> str:
    # NOTE: This list may be incorrect.
    return {
        0x01: "ボス",  # ウシまつ
        0x02: "ボス",  # おおなまず
        0x04: "1の巻",  # どくまんじゅう
        0x05: "1の巻",  # レッドスネーク
        0x06: "1の巻",  # しろぼうず
        0x07: "1の巻",  # カマおとこ
        0x08: "ボス",  # まむしおとこ
        0x09: "1の巻",  # フライングバム
        0x0A: "1の巻",  # スカルバット
        0x0B: "1の巻",  # ダイキチ
        0x0C: "1の巻",  # へルラッツ
        0x0D: "1の巻",  # アッカンべー
        0x0E: "1の巻",  # へらへら
        0x0F: "1の巻",  # クモジン
        0x10: "1の巻",  # からくりマン
        0x11: "1の巻",  # ブラックマン
        0x12: "1の巻",  # かげにん
        0x13: "1の巻",  # にんけん
        0x14: "ボス",  # デビルクローン
        0x15: "2の巻",  # べムガー
        0x16: "2の巻",  # デスグりーン
        0x17: "2の巻",  # へルモンキー
        0x18: "2の巻",  # シェルビー
        0x19: "2の巻",  # へびおんな
        0x1A: "2の巻",  # ドラゴンマン
        0x1B: "2の巻",  # ぎょろん
        0x1C: "2の巻",  # メカタツノコ
        0x1D: "ボス",  # タツノコつかい
        0x1E: "ボス",  # うつぼうず
        0x1F: "3の巻",  # おおとげむし
        0x20: "3の巻",  # マグマンゼりー
        0x21: "3の巻",  # ひふきガメ
        0x22: "3の巻",  # べロべロべー
        0x23: "3の巻",  # ヒダシメ
        0x24: "3の巻",  # バットクルス
        0x25: "3の巻",  # かえんマン
        0x26: "3の巻",  # いんねび
        0x27: "ボス",  # かえんだいおう
        0x28: "4の巻",  # マッドボアー
        0x29: "4の巻",  # ひゃっかんいぬ
        0x2A: "4の巻",  # ウルフマン
        0x2B: "4の巻",  # りトルエイプ
        0x2C: "4の巻",  # しろやまた
        0x2D: "4の巻",  # へルバット
        0x2E: "4の巻",  # くらやみマン
        0x2F: "4の巻",  # シャドウマン
        0x30: "ボス",  # ムササビだゆう
        0x31: "5の巻",  # あまのじゃく
        0x32: "5の巻",  # クレイジーカウ
        0x33: "5の巻",  # レッドソーサー
        0x34: "5の巻",  # しにがみこぞう
        0x35: "5の巻",  # にんげんもどき
        0x36: "5の巻",  # バトルナイト
        0x37: "ボス",  # コウモりだゆう
        0x38: "6の巻",  # ララバイかめん
        0x39: "6の巻",  # スーパークロン
        0x3A: "6の巻",  # シャーぺイン
        0x3B: "6の巻",  # サソラム
        0x3C: "6の巻",  # コカーメン
        0x3D: "6の巻",  # カーメン
        0x3E: "6の巻",  # きがマン
        0x3F: "ボス",  # ビッグカンカン
        0x40: "6の巻",  # マッドミイラ
        0x41: "6の巻",  # ファラー
        0x42: "6の巻",  # スフインツク
        0x43: "ボス",  # ツタンだいおう
        0x44: "7の巻",  # パニュロン
        0x45: "7の巻",  # ガオウ
        0x46: "7の巻",  # アイスマン
        0x47: "7の巻",  # ひょうがんだん
        0x48: "7の巻",  # ガンテツゾンビ
        0x49: "7の巻",  # ゆきひめ
        0x4A: "7の巻",  # ダークへッド
        0x4B: "7の巻",  # ひょうけつマン
        0x4C: "7の巻",  # ひょうがコング
        0x4D: "ボス",  # だるまだいし
        0x4E: "8の巻",  # へルファイヤー
        0x4F: "8の巻",  # どろたぼう
        0x50: "8の巻",  # ゾンビー
        0x51: "8の巻",  # のろいひめ
        0x52: "8の巻",  # ゾンビコウモり
        0x53: "8の巻",  # メタルパラソル
        0x54: "8の巻",  # あしがるゾンビ
        0x55: "8の巻",  # のろいマン
        0x56: "8の巻",  # スカルホッパー
        0x57: "8の巻",  # ミステりーアイ
        0x58: "ボス",  # ゾンビまおう
        0x59: "8の巻",  # かねくいだま
        0x5A: "8の巻",  # ラーゴン
        0x5B: "8の巻",  # モスカルラ
        0x5C: "8の巻",  # ダンダン
        0x5D: "8の巻",  # キンゾー
        0x5E: "8の巻",  # ガキゾンビ
        0x5F: "8の巻",  # ブレインソーサ
        0x60: "ボス",  # ロボゴールド
        0x61: "1の巻",  # キノコング
        0x63: "9の巻",  # あかぼうず
        0x64: "9の巻",  # ダークネス
        0x65: "9の巻",  # まぼろしかめん
        0x66: "9の巻",  # ろくろ
        0x67: "9の巻",  # まそうりょ
        0x68: "9の巻",  # はんにゃ
        0x69: "9の巻",  # ドラゴルド
        0x6A: "ボス",  # バイオフラワー
        0x6B: "9の巻",  # マンイーター
        0x6C: "9の巻",  # フラワー
        0x6D: "10の巻",  # みらいマン
        0x6E: "10の巻",  # ジョックー
        0x6F: "10の巻",  # ミンミン
        0x70: "10の巻",  # アイアンアイ
        0x71: "10の巻",  # ガンダーロボ
        0x72: "10の巻",  # レイザータンク
        0x73: "ボス",  # ボスガンダー１
        0x74: "ボス",  # ボスガンダー２
        0x75: "10の巻",  # メガべルガー
        0x76: "10の巻",  # レガルゴ
        0x77: "10の巻",  # スカイキラー
        0x78: "ボス",  # エイりアンドー
        0x79: "ボス",  # キラーウルフ
        0x7A: "ボス",  # メタルブロック
        0x7B: "ボス",  # フライウイドウ
        0x7C: "ボス",  # サーべンラガー
        0x7D: "ボス",  # マインマスター
        0x7E: "オニガランド",  # クレイジーババ
        0x7F: "オニガランド",  # まへいもち
        0x80: "オニガランド",  # ドグウアーマー
        0x81: "オニガランド",  # へビオトコ
        0x82: "ボス",  # イヌゾンビ
        0x83: "オニガランド",  # あおぼうず
        0x84: "オニガランド",  # オニデーモン
        0x85: "オニガランド",  # マグマン
        0x86: "オニガランド",  # モンスタージジ
        0x87: "オニガランド",  # あおきし
        0x88: "オニガランド",  # オニタコン
        0x89: "ボス",  # サルボス
        0x8A: "オニガランド",  # シーサーぺント
        0x8B: "オニガランド",  # ブルーアンクル
        0x8C: "オニガランド",  # スネークポッド
        0x8D: "オニガランド",  # レッドドッグ
        0x8E: "オニガランド",  # ブルードッグ
        0x8F: "オニガランド",  # バンコパ
        0x90: "オニガランド",  # ピーチボーイズ
        0x91: "ボス",  # モモタロゾンビ
        0x92: "ボス",  # キラーウルフ
        0x93: "オーロラ王国",  # グりーンアイ
        0x94: "オーロラ王国",  # つるりん
        0x95: "オーロラ王国",  # へビーガル
        0x96: "オーロラ王国",  # スカイマン
        0x97: "オーロラ王国",  # クレバス
        0x98: "オーロラ王国",  # グりーンケルプ
        0x99: "オーロラ王国",  # イノクラッシュ
        0x9A: "オーロラ王国",  # アイスファイヤ
        0x9B: "オーロラ王国",  # べムカッター
        0x9C: "オーロラ王国",  # がいこつむし
        0x9D: "オーロラ王国",  # シャドーマスク
        0x9E: "オーロラ王国",  # カニモンス
        0x9F: "オーロラ王国",  # アイスストーン
        0xA0: "オーロラ王国",  # ゆきみアイス
        0xA1: "オーロラ王国",  # ブルーザウルス
        0xA2: "オーロラ王国",  # じんめんいわ
        0xA3: "オーロラ王国",  # エレキラドン
        0xA4: "オーロラ王国",  # ブビ
        0xA5: "オーロラ王国",  # キルスライダー
        0xA6: "ボス",  # メタルブロック
        0xA9: "2の巻",  # しきゃくマン
        0xAA: "ボス",  # キンタロゾンビ
        0xAC: "ボス",  # ウシまつ
    }.get(enemy_id, "不明")


def _fill_enemy(prg_rom_bytes: bytes, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    row_index = 1
    header_row = [
        _ColumnStyle(caption="敵ID", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="敵名前"),
        _ColumnStyle(caption="初出"),
        _ColumnStyle(caption="基礎命", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="最小命", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="最大命", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="超力", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="攻撃", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="防御", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="スピード", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="経験値", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="獲得金", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="逃走"),
        _ColumnStyle(caption="行動回数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="回避補正", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="ミハギトきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="クルスきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="ベトきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="地震系超力きく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="火炎系超力きく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="水撃系・氷結系超力きく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="電撃系超力きく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="爆発系超力きく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="マフウジきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="マフウジ有効ターン数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="ララバイきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="ララバイ有効ターン数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="パラパきく率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="パラパ有効ターン数", cell_alignment=_RIGHT_ALIGNMENT),
    ]
    for i in range(_MAX_ACTIONS_PER_ENEMY):
        header_row.append(_ColumnStyle(caption=f"敵行動{i + 1}"))
        header_row.append(_ColumnStyle(caption=f"敵行動{i + 1}_確率", cell_alignment=_RIGHT_ALIGNMENT))
    header_row.append(_ColumnStyle(caption="ドロップアイテム1"))
    header_row.append(_ColumnStyle(caption="ドロップアイテム2"))
    header_row.append(_ColumnStyle(caption="ドロップアイテム確率", cell_alignment=_RIGHT_ALIGNMENT))
    header_row.append(_ColumnStyle(caption="備考"))
    _fill_worksheet_header_row(worksheet, row_index, header_row)
    worksheet.freeze_panes = "C2"
    choriki_endurance = {
        0: "100%",
        1: "70%",
        2: "30%",
        3: "0%",
    }
    debuff_endurance = {
        0: "100%",
        1: "70%",
        2: "30%",
        3: "0%",
    }
    tern_bucket = {
        0: "2 or 3",
        1: "4 or 5",
        2: "5 or 6",
    }
    escapable = {
        True: "可能",
        False: "不可",
    }
    attack_twice = {
        True: "1 or 2",
        False: "1",
    }
    hittability = {
        0: 0xC0,
        1: 0xA0,
        2: 0x80,
        3: 0,
    }
    mihagito_endurance = {
        True: "0%",
        False: "40%",
    }
    kurusu_endurance = {
        True: "0%",
        False: "20%",
    }
    beto_endurance = {
        True: "0%",
        False: "40%",
    }
    note_by_enemy_id = {
        0x78: "プレイヤーの行動選択後のターン開始時に命の上位バイトが#$7Fに上書きされる。(ターン開始時に命が32513(#$7F01)以上に回復する)",  # エイりアンドー
    }
    rows = []
    for enemy_id in range(1, _ENEMY_ID_COUNT + 1):
        enemy = _get_enemy(prg_rom_bytes, enemy_id)
        enemy_name = _get_enemy_name(prg_rom_bytes, enemy_id)
        mahuuji_endurance = debuff_endurance[enemy.mahuuji_endurance]
        if enemy.mahuuji_endurance == 3:
            assert enemy.mahuuji_effectiveness == 0
            mahuuji_tern_count = "-"
        else:
            mahuuji_tern_count = tern_bucket[enemy.mahuuji_effectiveness]
        lullaby_endurance = debuff_endurance[enemy.lullaby_endurance]
        if enemy.lullaby_endurance == 3:
            assert enemy.lullaby_effectiveness == 0
            lullaby_tern_count = "-"
        else:
            lullaby_tern_count = tern_bucket[enemy.lullaby_effectiveness]
        parapa_endurance = debuff_endurance[enemy.parapa_endurance]
        if enemy.parapa_endurance == 3:
            assert enemy.parapa_effectiveness == 0
            parapa_tern_count = "-"
        else:
            parapa_tern_count = tern_bucket[enemy.parapa_effectiveness]
        row = [
            enemy_id,
            enemy_name.enemy_name.strip(),
            _get_chapter_name_of_enemy(enemy_id),
            enemy.hp,
            enemy.min_hp,
            enemy.max_hp,
            enemy.cp,
            enemy.attack,
            enemy.defense,
            enemy.speed,
            enemy.money,
            enemy.experience,
            escapable[enemy.escapable],
            attack_twice[enemy.attack_twice],
            hittability[enemy.hittability],
            mihagito_endurance[enemy.mihagito_endurance],
            kurusu_endurance[enemy.kurusu_endurance],
            beto_endurance[enemy.beto_endurance],
            choriki_endurance[enemy.choriki_endurance_1],
            choriki_endurance[enemy.choriki_endurance_2],
            choriki_endurance[enemy.choriki_endurance_3_4],
            choriki_endurance[enemy.choriki_endurance_5],
            choriki_endurance[enemy.choriki_endurance_6],
            mahuuji_endurance,
            mahuuji_tern_count,
            lullaby_endurance,
            lullaby_tern_count,
            parapa_endurance,
            parapa_tern_count,
        ]
        enemy_action_pattern = _get_enemy_action_pattern(prg_rom_bytes, enemy.action_pattern_id)
        action_threshold_by_action_id = _aggregate_enemy_action_pattern(enemy_action_pattern)
        sorted_actions = sorted(action_threshold_by_action_id.items(), key=operator.itemgetter(1), reverse=True)
        for i in range(_MAX_ACTIONS_PER_ENEMY):
            if i < len(sorted_actions):
                (action_id, threshold) = sorted_actions[i]
                action_name = _get_action_name(prg_rom_bytes, action_id).action_name.strip()
                row.append(action_name)
                row.append(f"{round(threshold / (0x100 * 2) * 100, 2):.02f}% ({threshold} / {0x100 * 2})")
            else:
                row.append("-")
                row.append("-")
        item_drop_pattern = _get_item_drop_pattern(prg_rom_bytes, enemy.item_drop_pattern_id)
        has_drop_item = False
        if item_drop_pattern.drop_item_id_0 == 0:
            row.append("-")
        else:
            has_drop_item = True
            row.append(_get_item_name(prg_rom_bytes, item_drop_pattern.drop_item_id_0).item_name.strip())
        if item_drop_pattern.drop_item_id_1 == 0:
            assert not has_drop_item
            row.append("-")
        else:
            assert has_drop_item
            has_drop_item = True
            row.append(_get_item_name(prg_rom_bytes, item_drop_pattern.drop_item_id_1).item_name.strip())
        if not has_drop_item:
            row.append("-")
        elif item_drop_pattern.item_drop_threshold == 0:
            row.append("100.00%")
        else:
            row.append(f"{round(item_drop_pattern.item_drop_threshold / 0x100 * 100, 2):.02f}% ({item_drop_pattern.item_drop_threshold} / {0x100})")
        row.append(note_by_enemy_id.get(enemy_id, "-"))
        rows.append(row)

    def _get_sort_key(row: List[Any]) -> Tuple[int, int]:
        enemy_id = row[0]
        enemy_chapter_priority = [
            "1の巻",
            "2の巻",
            "3の巻",
            "4の巻",
            "5の巻",
            "6の巻",
            "7の巻",
            "8の巻",
            "9の巻",
            "10の巻",
            "オニガランド",
            "オーロラ王国",
            "ボス",
            "不明",
        ].index(row[2])
        return (enemy_chapter_priority, enemy_id)

    for row in sorted(rows, key=_get_sort_key):
        row_index += 1
        _fill_worksheet_row(worksheet, row_index, header_row, row)


def _get_encounter_threshold(prg_rom_bytes: bytes, encounter_threshold_id: int) -> int:
    assert 0 <= encounter_threshold_id <= 7
    if encounter_threshold_id == 0:
        return 0
    return prg_rom_bytes[0x00FD7D + (encounter_threshold_id - 1)]


def _fill_map(prg_rom_bytes: bytes, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    description_by_map_id = {
        0x0000: "ワールドマップ",
        0x0001: "オニガランドワールドマップ",
        0x0002: "オーロラ王国ワールドマップ",
        0x0004: "柳生の里",
        0x0008: "尾張の町:奉行所1階",
        0x0009: "尾張の町:奉行所地下牢",
        0x0025: "アザラシ村とゆきおとこ村の間の洞窟",
        0x0026: "オーロラ村とアザラシ村の間の洞窟2",
        0x0028: "浪速の都:越後屋への地下通路(左から右)",
        0x002A: "なまず大明神:入り口の社",
        0x002B: "なまず大明神:おおなまずのフロア",
        0x002C: "からくり城:1階",
        0x002E: "からくり城:地下1階牢屋",
        0x0030: "柳生の里(江戸の町崩壊後)",
        0x0031: "からくり城:入り口の細い通路",
        0x0035: "柳生の里:但馬邸(江戸の町崩壊後)",
        0x003A: "シードラゴン:縦に長い階段があるフロア",
        0x003B: "シードラゴン:最下層(水の流れているフロア)",
        0x003C: "シードラゴン:一番上の層の右の部屋・3番目の層の右の部屋(何も無い部屋)",
        0x003D: "シードラゴン:乙姫の前の人が7人いる部屋",
        0x003E: "シードラゴン:たつのこつかいのフロアの1つ前の宝箱が6個ある部屋",
        0x003F: "シードラゴン:うつぼうずのフロア",
        0x0040: "龍宮からシードラゴンへの通路1",
        0x0041: "シードラゴン:入ってすぐのフロア",
        0x0042: "シードラゴン:最下層からうつぼうずの途中の部屋1",
        0x0043: "龍宮",
        0x0044: "シードラゴン:たつのこつかいのいる部屋",
        0x0045: "シードラゴン:一番上の層の左の部屋・2番目の層の左の部屋(たるが2個ある部屋)",
        0x0047: "シードラゴン:2番目の層の真ん中の部屋・3番目の層の左の部屋(たるが5個ある部屋)",
        0x0048: "シードラゴン:2番目の層の右の部屋(うつぼのカギがある部屋)",
        0x004A: "龍宮からシードラゴンへの通路2",
        0x004C: "シードラゴン:最下層からうつぼうずの途中の部屋2",
        0x004D: "シードラゴン:最下層からうつぼうずの途中の部屋3",
        0x004E: "シードラゴン:最下層からうつぼうずの途中の部屋4",
        0x004F: "シードラゴン:最下層からうつぼうずの途中の部屋5",
        0x0050: "火炎城:フロア1(火炎城最初のフロア)",
        0x0051: "暗闇城:地下4階(こうもりだゆうのフロア)",
        0x0053: "龍の祠(さばのすけで龍宮の入り口を見つけるフロア)",
        0x0054: "のろい城:2階",
        0x0056: "火炎城:フロア2",
        0x0058: "火炎城:フロア3",
        0x0059: "火炎城:フロア4(横1列のフロア)",
        0x005A: "氷結城:左の塔4階",
        0x005B: "氷結城:右の塔4階",
        0x005C: "イワンのだっしゅつ",
        0x005D: "龍宮への入り口の次の真っ黒のフロア(じゅうべえ落下)",
        0x005F: "シードラゴン:乙姫の部屋",
        0x0060: "火炎城:フロア5",
        0x0061: "暗闇城:地下1階(入ってすぐのフロア)",
        0x0062: "暗闇城:地下2階",
        0x0063: "暗闇城:地下3階(牢屋)",
        0x0064: "暗闇城:地下2階(ヘルバットのフロア)",
        0x0065: "浪速の都(からくり城攻略前の暗い状態)",
        0x0067: "浪速の都:越後屋からからくり城への地下通路(下から上)",
        0x0068: "飢餓城:1階",
        0x006B: "浪速の都:越後屋",
        0x006E: "飢餓城:2階",
        0x006F: "飢餓城:3階(スフインツクのフロア)",
        0x0070: "飢餓城:地下1階",
        0x0071: "飢餓城:3階",
        0x0072: "飢餓城:3階(ビッグカンカンからスフインツクの間の通路)",
        0x0073: "飢餓城:4階(ツタンだいおうのフロア)",
        0x0074: "飢餓城:ビッグカンカンのフロア",
        0x0076: "氷結城:左の塔2階",
        0x0077: "氷結城:左の塔3階",
        0x007B: "氷結城:5階(メガトンコインを持っていると落ちてしまうところ)",
        0x007C: "氷結城:だるまたいしのいるフロア",
        0x007D: "氷結城:右の塔1階",
        0x007E: "氷結城:右の塔3階",
        0x0083: "かぶとがに大明神:入り口の社",
        0x0084: "かぶとがに大明神:内部",
        0x0085: "安芸の町",
        0x0086: "伊予の町",
        0x0088: "伊予の町から土佐の町への地下道",
        0x0089: "宇宙(タコリアンのUFOでの移動画面)",
        0x008A: "つちのこ大明神",
        0x008B: "土佐の町(シードラゴン攻略前)",
        0x008D: "のろい城:地下1階",
        0x008E: "のろい城:1階(入ってすぐのフロア)",
        0x008F: "のろい城:1階",
        0x0090: "のろい城:3階(棺桶が多いフロア)・4階",
        0x0091: "のろい城:5階(ゾンビまおうのフロア)",
        0x0092: "のろい城:3階(牢屋のあるフロア)",
        0x0094: "のろい城:ゾンビマシン",
        0x0097: "岬の小屋(さばのすけのいるフロア)",
        0x0098: "呉別府の渡し",
        0x0099: "異人の町",
        0x009A: "隼人の渡し",
        0x00A2: "モンゴレンの町:あおいほんがあるフロア",
        0x00A5: "オーロラ村とアザラシ村の間の洞窟1",
        0x00A6: "黄金洞窟",
        0x00A8: "ゆきおとこ村北のガンちゃんで岩を退けるフロア",
        0x00A9: "浪速の都:越後屋の隠し通路部屋",
        0x00AC: "未来城:左の塔上層1階",
        0x00AD: "未来城:右の塔上層(最下層へ落下させられるフロア)",
        0x00AF: "隠れ湯",
        0x00B0: "富士山への地下通路",
        0x00B8: "未来城:パームロケットが貰える部屋",
        0x00BD: "ミロクの洞窟",
        0x00BE: "未来城:外観(入ってすぐのフロア)",
        0x00C0: "未来城:格納庫(床下パネルを調べながら進むフロアの途中にある上下に入り口のある部屋)",
        0x00C4: "未来城:左の塔下層(左の塔入ってすぐのフロア)",
        0x00C5: "未来城:左の塔最上階(ボスガンダー1のフロア)",
        0x00C9: "未来城:中央の橋(パームロケットを使うフロア)",
        0x00CA: "未来城:右の塔(上層から最下層への落下画面)",
        0x00CB: "未来城:右の塔最下層(ボスガンダー2のフロア)",
        0x00CF: "未来城:格納庫(ドールのカギを使い入ってすぐのフロア)",
        0x00D6: "未来城:格納庫(床下パネルを調べながら進むフロア)",
        0x00DC: "未来城:中央の塔(マインマスターのフロア)",
        0x00E0: "未来城:中央の塔(マインマスターのフロアのひとつ前のフロア)",
        0x00E1: "モモタロゾンビの城1階",
        0x00E2: "モモタロゾンビの城2階",
        0x00E3: "モモタロゾンビの城3階",
        0x00E5: "オーロラ村:王宮",
        0x00E6: "未来城:左の塔上層2階",
        0x00E7: "未来城:左の塔上層3階",
        0x00E8: "未来城:左の塔上層4階",
        0x00E9: "未来城:左の塔上層5階",
        0x00EA: "未来城:右の塔最上階",
        0x00EB: "未来城:右の塔最上階-1階",
        0x00EC: "ゆきおとこ村からクーラーの洞窟の間の洞窟",
        0x00ED: "ゆきおとこ村",
        0x00EE: "北の洞窟:入ってすぐのフロア",
        0x00EF: "北の洞窟:2番目のフロア",
        0x00F0: "オーロラ村",
        0x00F1: "モモタロゾンビの城4階",
        0x00F2: "モモタロゾンビの城5階",
        0x00F3: "クーラーの洞窟:入ってすぐのフロア",
        0x00F4: "クーラーの洞窟:メタルブロックのいるフロア",
        0x00F5: "隠れ湯:みかづきの部屋",
        0x00F6: "薩摩の町(火炎城攻略前)",
        0x00F7: "薩摩の町(かえんだいおう戦後)",
        0x00F8: "薩摩の町(かえんだいおう潜伏時)",
        0x00FC: "安芸の町:かごちゃんの部屋",
        0x00FD: "琉球の村",
        0x00FE: "琉球の村から火炎城への地下道",
        0x0100: "屋久島",
        0x0101: "屋久島:杉の子大明神",
        0x0102: "壱岐(天狗のいるフロア)",
        0x0103: "オロ島(シロのいるフロア)",
        0x0104: "門司の村",
        0x0105: "門司の村:鬼の涙を使う穴がある部屋",
        0x0106: "門司の村:鬼の涙を使うフロア",
        0x0107: "下関の村",
        0x010A: "黄泉の洞窟:入ってすぐのフロア",
        0x010C: "黄泉の洞窟:エイリアンドール跡地",
        0x010D: "カムカムの渡し",
        0x010E: "黄泉の洞窟:タコリアンのいるフロア",
        0x010F: "プーサンの村",
        0x0110: "暗闇城から黄泉の洞窟への地下通路",
        0x0111: "異星の廃都",
        0x0112: "長門の村",
        0x0113: "コンコンの町",
        0x0116: "ホルクロア:まがつたまがあるフロア",
        0x0117: "白ウサギ大明神",
        0x0118: "ジャンパイの町",
        0x011A: "ポキン",
        0x011C: "クーロン城:入ってすぐのフロア",
        0x011D: "クーロン城:地下牢",
        0x011E: "クーロン城:コスモトロンがあるフロア",
        0x011F: "ソウレンの村",
        0x0121: "シーサンプータ",
        0x0122: "三里の長城",
        0x0123: "モンゴレンの町",
        0x0125: "ハルビンタの村",
        0x0127: "からくり城(ワールドマップから入った場合;入り口が無い)",
        0x0128: "コウモリ洞窟(しんかげがあるフロア)",
        0x012A: "ウラジョスト",
        0x012C: "ババロフの町",
        0x012D: "青い石碑(ノルンのなみだを使うフロア)",
        0x012E: "最果ての洞窟(イワンの埋まっているフロア)",
        0x012F: "石狩の町",
        0x0130: "北の神々の祠",
        0x0131: "まりもの里",
        0x0133: "函館の村",
        0x0134: "りんご村",
        0x0135: "イタコ村",
        0x0137: "十和田の石碑",
        0x0138: "なんぶの町",
        0x013C: "氷結城:入り口",
        0x013D: "のろい城:入り口",
        0x013E: "あきんどタウン",
        0x013F: "ミミズク大明神",
        0x0140: "いけない渡し",
        0x0142: "あわの村",
        0x0143: "江戸の町",
        0x0146: "エンディング:マインマスター戦後のフロア",
        0x0147: "シバレンの村",
        0x0148: "のろい城:5階からゾンビマシンの間の移動",
        0x014D: "千里の長城",
        0x014E: "富士山",
        0x014F: "ホルクロア",
        0x0150: "薩摩の町の右下の火山",
        0x0152: "しろくま村",
        0x0156: "アザラシ村",
        0x0158: "シロ編エンディングのスギ",
        0x0159: "北の洞窟:ガンちゃんがいるフロア",
        0x015A: "ミミナリ島の祠1階",
        0x015B: "ミミナリ島の祠2階(キンタロゾンビのフロア)",
        0x015C: "トンカチ島の祠(チューリップ)",
        0x015D: "トンカチ島の祠(チューリップで転送後)",
    }
    row_index = 1
    header_row = [
        _ColumnStyle(caption="マップID", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="マップ説明"),
        _ColumnStyle(caption="敵グループパターンリストID", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="敵エンカウント確率", cell_alignment=_RIGHT_ALIGNMENT),
    ]
    _fill_worksheet_header_row(worksheet, row_index, header_row)
    worksheet.freeze_panes = "B2"
    for map_id in range(0, _MAP_ID_COUNT):
        map_data = _get_map(prg_rom_bytes, map_id)
        if map_id in (0x00, 0x01, 0x02):
            # 0x00: ワールドマップ
            # 0x01: オニガランドワールドマップ
            # 0x02: オーロラ王国ワールドマップ
            encounter_rate = "移動先のマスの種類により変わる"
        else:
            encounter_threshold = _get_encounter_threshold(prg_rom_bytes, map_data.encounter_threshold_id)
            encounter_rate = f"{round(encounter_threshold / 0x100 * 100, 2):.02f}% ({encounter_threshold} / {0x100})"
        row = [map_id, description_by_map_id.get(map_id, ""), map_data.enemy_group_pattern_list_id, encounter_rate]
        row_index += 1
        _fill_worksheet_row(worksheet, row_index, header_row, row)


def _aggregate_enemy_action_pattern_ids(enemy_group_pattern_ids: Sequence[int]) -> Sequence[Tuple[int, int]]:
    return tuple((enemy_group_pattern_id, len(tuple(ids))) for enemy_group_pattern_id, ids in itertools.groupby(sorted(enemy_group_pattern_ids)))


def _fill_enemy_group_pattern_list(prg_rom_bytes: bytes, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    row_index = 1
    header_row = [
        _ColumnStyle(caption="敵グループパターンリストID", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="敵グループパターンID", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="敵グループパターンリスト内確率", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="敵合計数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="グループ数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="グループ1敵名前"),
        _ColumnStyle(caption="グループ1敵数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="グループ2敵名前"),
        _ColumnStyle(caption="グループ2敵数", cell_alignment=_RIGHT_ALIGNMENT),
        _ColumnStyle(caption="グループ3敵名前"),
        _ColumnStyle(caption="グループ3敵数", cell_alignment=_RIGHT_ALIGNMENT),
    ]
    _fill_worksheet_header_row(worksheet, row_index, header_row)
    worksheet.freeze_panes = "B2"
    for enemy_group_pattern_list_id in range(0, _ENEMY_GROUP_PATTERN_LIST_ID_COUNT):
        enemy_group_pattern_list = _get_enemy_group_pattern_list(prg_rom_bytes, enemy_group_pattern_list_id)
        enemy_group_pattern_ids_count = _aggregate_enemy_action_pattern_ids(enemy_group_pattern_list.enemy_group_pattern_ids)
        sum_enemy_group_pattern_ids_count = sum(item[1] for item in enemy_group_pattern_ids_count)
        if sum_enemy_group_pattern_ids_count:
            for enemy_group_pattern_id, count in enemy_group_pattern_ids_count:
                row: List[Union[int, str]] = [
                    enemy_group_pattern_list_id,
                    enemy_group_pattern_id,
                ]
                rate = f"{round(count / sum_enemy_group_pattern_ids_count * 100, 2):.02f}% ({count} / {sum_enemy_group_pattern_ids_count})"
                row.append(rate)
                enemy_group_pattern = _get_enemy_group_pattern(prg_rom_bytes, enemy_group_pattern_id)
                row.append(enemy_group_pattern.enemy_group_size)
                row.append(enemy_group_pattern.enemy_group_0_size + enemy_group_pattern.enemy_group_1_size + enemy_group_pattern.enemy_group_2_size)
                if enemy_group_pattern.enemy_group_0_enemy_id is not None:
                    enemy_0_name = _get_enemy_name(prg_rom_bytes, enemy_group_pattern.enemy_group_0_enemy_id).enemy_name.strip()
                    row.append(enemy_0_name)
                    row.append(enemy_group_pattern.enemy_group_0_size)
                else:
                    row.append("-")
                    row.append("-")
                if enemy_group_pattern.enemy_group_1_enemy_id is not None:
                    enemy_1_name = _get_enemy_name(prg_rom_bytes, enemy_group_pattern.enemy_group_1_enemy_id).enemy_name.strip()
                    row.append(enemy_1_name)
                    row.append(enemy_group_pattern.enemy_group_1_size)
                else:
                    row.append("-")
                    row.append("-")
                if enemy_group_pattern.enemy_group_2_enemy_id is not None:
                    enemy_2_name = _get_enemy_name(prg_rom_bytes, enemy_group_pattern.enemy_group_2_enemy_id).enemy_name.strip()
                    row.append(enemy_2_name)
                    row.append(enemy_group_pattern.enemy_group_2_size)
                else:
                    row.append("-")
                    row.append("-")
                row_index += 1
                _fill_worksheet_row(worksheet, row_index, header_row, row)
        else:
            # No group patterns in the list.
            row = [
                enemy_group_pattern_list_id,
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
            ]
            row_index += 1
            _fill_worksheet_row(worksheet, row_index, header_row, row)


def _fill_world_map(prg_rom_bytes: bytes, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    description_by_prg_rom_address = {
        0x00FCE7: "ワールドマップ(森(木2本))",
        0x00FCE8: "ワールドマップ(森(木1本))",
        0x00FCE9: "ワールドマップ(??)",
        0x00FCEA: "ワールドマップ(雪木(木1本))",
        0x00FCEB: "ワールドマップ(山(茶色))",
        0x00FCEC: "ワールドマップ(岩山(1マス;進入不可))",
        0x00FCED: "ワールドマップ(雪山(1マス;進入不可))",
        0x00FCEE: "ワールドマップ(平野(緑))",
        0x00FCEF: "ワールドマップ(橋(上から下))",
        0x00FCF0: "ワールドマップ(海(左側が橋の影;進入不可))",
        0x00FCF1: "ワールドマップ(橋(左から右))",
        0x00FCF2: "ワールドマップ(海(上側が橋の影;進入不可))",
        0x00FCF3: "ワールドマップ(沼(緑))",
        0x00FCF4: "ワールドマップ(水(岸なし;進入不可))",
        0x00FCF5: "ワールドマップ(水(上側が岸;進入不可))",
        0x00FCF6: "ワールドマップ(水(上側と左側が岸;進入不可))",
        0x00FCF7: "ワールドマップ(水(左側が岸;進入不可))",
        0x00FCF8: "ワールドマップ(??)",
        0x00FCF9: "ワールドマップ(水(下側が岸;進入不可))",
        0x00FCFA: "ワールドマップ(水(左側と下側が岸;進入不可))",
        0x00FCFB: "ワールドマップ(水(下側と右側が岸;進入不可))",
        0x00FCFC: "ワールドマップ(??)",
        0x00FCFD: "ワールドマップ(水(左側が岸;進入不可))",
        0x00FCFE: "ワールドマップ(水(右側が岸;進入不可))",
        0x00FCFF: "ワールドマップ(水(上側と下側が岸;進入不可))",
        0x00FD00: "ワールドマップ(水(左側と右側が岸;進入不可))",
        0x00FD01: "ワールドマップ(平野(緑;下側が岸))",
        0x00FD02: "ワールドマップ(砂漠)",
        0x00FD03: "ワールドマップ(??)",
        0x00FD04: "ワールドマップ(ヤシの木)",
        0x00FD05: "ワールドマップ(??)",
        0x00FD06: "ワールドマップ(城壁(進入不可))",
        0x00FD07: "ワールドマップ(スギ(2マスの上))",
        0x00FD08: "ワールドマップ(雪原)",
        0x00FD09: "ワールドマップ(スギ(2マスの下))",
        0x00FD0A: "ワールドマップ(雪原(下側が岸))",
        0x00FD0B: "ワールドマップ(岩山(2マスの左;進入不可))",
        0x00FD0C: "ワールドマップ(岩山(2マスの右;進入不可))",
        0x00FD0D: "ワールドマップ(町(2マス;左))",
        0x00FD0E: "ワールドマップ(町(2マス;右))",
        0x00FD0F: "ワールドマップ(洞窟(はしご))",
        0x00FD10: "ワールドマップ(鳥居)",
        0x00FD11: "ワールドマップ(町(1マス))",
        0x00FD12: "ワールドマップ(氷結城(4マス;左上))",
        0x00FD13: "ワールドマップ(氷結城(4マス;右上))",
        0x00FD14: "ワールドマップ(氷結城(4マス;左下))",
        0x00FD15: "ワールドマップ(氷結城(4マス;右下))",
        0x00FD16: "ワールドマップ(火山・ピラミッド・呪い城(4マス;左上))",
        0x00FD17: "ワールドマップ(火山・ピラミッド・呪い城(4マス;右上))",
        0x00FD18: "ワールドマップ(火山・ピラミッド・呪い城(4マス;左下))",
        0x00FD19: "ワールドマップ(火山・ピラミッド・呪い城(4マス;右下))",
        0x00FD1A: "ワールドマップ(町跡地)",
        0x00FD1B: "オニガランド(森(木1本))",
        0x00FD1C: "オニガランド(??)",
        0x00FD1D: "オニガランド(平野(茶))",
        0x00FD1E: "オニガランド(森(幹無し))",
        0x00FD1F: "オニガランド(森(木2本))",
        0x00FD20: "オニガランド(平野(緑))",
        0x00FD21: "オニガランド(??)",
        0x00FD22: "オニガランド(??)",
        0x00FD23: "オニガランド(??)",
        0x00FD24: "オニガランド(??)",
        0x00FD25: "オニガランド(??)",
        0x00FD26: "オニガランド(水(岸なし))",
        0x00FD27: "オニガランド(水(上側が岸))",
        0x00FD28: "オニガランド(水(上側と左側が岸))",
        0x00FD29: "オニガランド(水(上側と右側が岸))",
        0x00FD2A: "オニガランド(??)",
        0x00FD2B: "オニガランド(水(下側が岸))",
        0x00FD2C: "オニガランド(水(左側と下側が岸))",
        0x00FD2D: "オニガランド(水(下側と右側が岸))",
        0x00FD2E: "オニガランド(??)",
        0x00FD2F: "オニガランド(水(左側が岸))",
        0x00FD30: "オニガランド(水(右側が岸))",
        0x00FD31: "オニガランド(水(上側と下側が岸))",
        0x00FD32: "オニガランド(水(左側と右側が岸))",
        0x00FD33: "オニガランド(平野(茶;下側が岸))",
        0x00FD34: "オニガランド(トゲ)",
        0x00FD35: "オニガランド(溶岩)",
        0x00FD36: "オニガランド(??)",
        0x00FD37: "オニガランド(??)",
        0x00FD38: "オニガランド(??)",
        0x00FD39: "オニガランド(??)",
        0x00FD3A: "オニガランド(??)",
        0x00FD3B: "オニガランド(??)",
        0x00FD3C: "オニガランド(??)",
        0x00FD3D: "オニガランド(??)",
        0x00FD3E: "オニガランド(??)",
        0x00FD3F: "オニガランド(??)",
        0x00FD40: "オニガランド(??)",
        0x00FD41: "オニガランド(??)",
        0x00FD42: "オニガランド(??)",
        0x00FD43: "オニガランド(??)",
        0x00FD44: "オニガランド(??)",
        0x00FD45: "オニガランド(??)",
        0x00FD46: "オニガランド(??)",
        0x00FD47: "オニガランド(??)",
        0x00FD48: "オニガランド(??)",
        0x00FD49: "オニガランド(??)",
        0x00FD4A: "オニガランド(??)",
        0x00FD4B: "オニガランド(??)",
        0x00FD4C: "オニガランド(??)",
        0x00FD4D: "オニガランド(??)",
        0x00FD4E: "オーロラ王国(雪原)",
        0x00FD4F: "オーロラ王国(雪原(下側が岸))",
        0x00FD50: "オーロラ王国(??)",
        0x00FD51: "オーロラ王国(??)",
        0x00FD52: "オーロラ王国(??)",
        0x00FD53: "オーロラ王国(??)",
        0x00FD54: "オーロラ王国(土)",
        0x00FD55: "オーロラ王国(??)",
        0x00FD56: "オーロラ王国(??)",
        0x00FD57: "オーロラ王国(??)",
        0x00FD58: "オーロラ王国(??)",
        0x00FD59: "オーロラ王国(??)",
        0x00FD5A: "オーロラ王国(??)",
        0x00FD5B: "オーロラ王国(??)",
        0x00FD5C: "オーロラ王国(??)",
        0x00FD5D: "オーロラ王国(??)",
        0x00FD5E: "オーロラ王国(??)",
        0x00FD5F: "オーロラ王国(??)",
        0x00FD60: "オーロラ王国(??)",
        0x00FD61: "オーロラ王国(??)",
        0x00FD62: "オーロラ王国(??)",
        0x00FD63: "オーロラ王国(??)",
        0x00FD64: "オーロラ王国(土(下側が岸))",
        0x00FD65: "オーロラ王国(??)",
        0x00FD66: "オーロラ王国(??)",
        0x00FD67: "オーロラ王国(階段)",
        0x00FD68: "オーロラ王国(??)",
        0x00FD69: "オーロラ王国(??)",
        0x00FD6A: "オーロラ王国(??)",
        0x00FD6B: "オーロラ王国(??)",
        0x00FD6C: "オーロラ王国(??)",
        0x00FD6D: "オーロラ王国(??)",
        0x00FD6E: "オーロラ王国(??)",
        0x00FD6F: "オーロラ王国(??)",
        0x00FD70: "オーロラ王国(??)",
        0x00FD71: "オーロラ王国(??)",
        0x00FD72: "オーロラ王国(??)",
        0x00FD73: "オーロラ王国(??)",
        0x00FD74: "オーロラ王国(??)",
        0x00FD75: "オーロラ王国(??)",
        0x00FD76: "オーロラ王国(??)",
        0x00FD77: "オーロラ王国(??)",
        0x00FD78: "オーロラ王国(??)",
        0x00FD79: "オーロラ王国(??)",
        0x00FD7A: "オーロラ王国(??)",
        0x00FD7B: "オーロラ王国(??)",
        0x00FD7C: "オーロラ王国(??)",
    }
    row_index = 1
    header_row = [
        _ColumnStyle(caption="マップ説明・マス説明"),
        _ColumnStyle(caption="敵エンカウント確率", cell_alignment=_RIGHT_ALIGNMENT),
    ]
    _fill_worksheet_header_row(worksheet, row_index, header_row)
    worksheet.freeze_panes = "B2"
    for prg_rom_address in range(0x00FCE7, 0x00FD7C + 1):
        encounter_threshold_id = prg_rom_bytes[prg_rom_address]
        encounter_threshold = _get_encounter_threshold(prg_rom_bytes, encounter_threshold_id)
        encounter_rate = f"{round(encounter_threshold / 0x100 * 100, 2):.02f}% ({encounter_threshold} / {0x100})"
        row = [description_by_prg_rom_address[prg_rom_address], encounter_rate]
        row_index += 1
        _fill_worksheet_row(worksheet, row_index, header_row, row)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_ines_file_path", type=str)
    parser.add_argument("output_excel_file_path", type=str)
    args = parser.parse_args()

    prg_rom_bytes = _read_prg_rom(args.input_ines_file_path)
    expected_prg_rom_crc = 0x29C61B41
    if binascii.crc32(prg_rom_bytes) != expected_prg_rom_crc:
        raise ValueError("Unexpected PRG ROM CRC")

    workbook = openpyxl.Workbook()
    workbook.active.title = "味方キャラステータス"
    _fill_player_character(prg_rom_bytes, workbook.active)
    _fill_enemy(prg_rom_bytes, workbook.create_sheet("敵キャラステータス"))
    _fill_map(prg_rom_bytes, workbook.create_sheet("マップ"))
    _fill_world_map(prg_rom_bytes, workbook.create_sheet("ワールドマップ"))
    _fill_enemy_group_pattern_list(prg_rom_bytes, workbook.create_sheet("敵グループパターンリスト"))
    workbook.save(args.output_excel_file_path)


if __name__ == "__main__":
    main()
